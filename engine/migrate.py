# -*- coding: utf-8 -*-
"""Migrate master-sheet.xlsx into the operational StateStore.

Rules:
- back up the workbook before migration;
- preserve/increment the existing state version when --force is used;
- verify persisted source-row counts after the commit;
- never treat a missing/empty production store as a friendly default.
"""
from __future__ import annotations

import datetime as dt
import os
import shutil
import sys

BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from openpyxl import load_workbook
from store import DATA_DIR, Store, log_event

SHEET = os.path.join(BASE, "data", "master-sheet.xlsx")
PRE_MIGRATION_BACKUPS = os.path.join(DATA_DIR, "30-state-backups")

TAB_MAP = {
    "مهام": "tasks",
    "مشاريع": "projects",
    "عملاء وفرص": "leads",
    "مؤشرات القسم": "kpis",
    "مواعيد": "meetings",
    "قرارات": "decisions",
    "متابعة مرضى": "followups",
    "صندوق الصوت": "voice",
    "تعلم": "learning",
    "مالية": "finance",
}


def rows_of(wb, tab):
    ws = wb[tab]
    headers = [c.value for c in ws[1]]
    out = []
    for row_values in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None for v in row_values):
            continue
        row = {}
        for header, value in zip(headers, row_values):
            if isinstance(value, dt.datetime):
                value = value.date() if value.hour == value.minute == value.second == 0 else value
            row[header] = value
        out.append(row)
    return out


def backup_workbook():
    if not os.path.exists(SHEET):
        raise RuntimeError(f"master workbook missing: {SHEET}")
    os.makedirs(PRE_MIGRATION_BACKUPS, exist_ok=True)
    stamp = dt.datetime.now().strftime("%Y%m%d-%H%M%S")
    target = os.path.join(PRE_MIGRATION_BACKUPS, f"master-sheet-{stamp}.xlsx")
    shutil.copy2(SHEET, target)
    return target


def build_state():
    wb = load_workbook(SHEET, data_only=True)
    missing = [tab for tab in TAB_MAP if tab not in wb.sheetnames]
    if missing:
        raise RuntimeError("missing workbook tabs: " + ", ".join(missing))

    state = {"meta": {}, "waiting_for": [], "action_queue": [], "manager_markers": {}}
    for tab, key in TAB_MAP.items():
        state[key] = rows_of(wb, tab)

    for lead in state["leads"]:
        if lead.get("الحالة") == "انتظار رد":
            state["waiting_for"].append({
                "item": f"رد من {lead['الجهة']} بشأن {lead.get('الخدمة', '')}".strip(),
                "source": "عملاء وفرص",
                "expected_from": lead["الجهة"],
                "since": lead.get("آخر تواصل"),
                "follow_up_date": None,
            })
    for project in state["projects"]:
        if project.get("الحالة") == "انتظار":
            state["waiting_for"].append({
                "item": f"تحرّك في مشروع {project['المشروع']} ({project.get('الخطوة التالية', '')})".strip(),
                "source": "مشاريع",
                "expected_from": "داخلي",
                "since": project.get("آخر تقدم"),
                "follow_up_date": None,
            })
    return state


def main():
    store = Store()
    force = "--force" in sys.argv
    if os.path.exists(store.path) and not force:
        print("state.json موجود أصلًا. لإعادة الترحيل استخدم: python3 engine/migrate.py --force")
        raise SystemExit(0)

    backup = backup_workbook()
    state = build_state()
    expected_source_rows = sum(len(state[key]) for key in TAB_MAP.values())

    store = Store()
    store.commit(
        state,
        "migrate_from_sheet",
        rows={key: len(value) for key, value in state.items() if isinstance(value, list)},
        source_backup=backup,
    )
    store.reload()
    store.validate(require_nonempty=True)

    persisted_source_rows = sum(len(store.data.get(key, [])) for key in TAB_MAP.values())
    if persisted_source_rows != expected_source_rows:
        raise RuntimeError(
            f"migration row-count mismatch: source={expected_source_rows} persisted={persisted_source_rows}"
        )

    log_event(
        "migrate_done",
        source="master-sheet.xlsx",
        backup=backup,
        source_rows=expected_source_rows,
        state_version=store.data["meta"]["version"],
    )
    print(
        f"✅ رُحّل الشيت إلى {store.path} — {expected_source_rows} صفًا مصدرًا + "
        f"{len(store.data['waiting_for'])} عنصر انتظار مشتق."
    )
    print(f"✅ نسخة ما قبل الترحيل: {backup}")
    print(f"✅ State version: {store.data['meta']['version']}")


if __name__ == "__main__":
    main()
